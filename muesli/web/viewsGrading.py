# -*- coding: utf-8 -*-
#
# muesli/web/viewsGrading.py
#
# This file is part of MUESLI.
#
# Copyright (C) 2011, Matthias Kuemmerer <matthias (at) matthias-k.org>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
import base64

import numpy
from matplotlib import pyplot

from muesli import models
from muesli import utils
from muesli.parser import Parser
from muesli.web.context import *
from muesli.web.forms import *

from pyramid.view import view_config
from pyramid.response import Response
from pyramid.httpexceptions import HTTPNotFound, HTTPFound
from pyramid.url import route_url
from sqlalchemy.orm import exc
from sqlalchemy.sql import func
import sqlalchemy
from matplotlib.ticker import MaxNLocator
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile
from muesli.web.tooltips import grading_edit_tooltips

import re
import os
import io
import datetime

@view_config(route_name='grading_edit', renderer='muesli.web:templates/grading/edit.pt', context=GradingContext, permission='edit')
class Edit:
    def __init__(self, request):
        self.request = request
        self.db = self.request.db
    def __call__(self):
        grading = self.request.context.grading
        form = GradingEdit(self.request, grading)
        if self.request.method == 'POST' and form.processPostData(self.request.POST):
            form.saveValues()
            self.request.db.commit()
            form.message = "Änderungen gespeichert."
        return {'grading': grading,
                'form': form
               }

@view_config(route_name='grading_associate_exam', context=GradingContext, permission='edit')
class AssociateExam:
    def __init__(self, request):
        self.request = request
        self.db = self.request.db
    def __call__(self):
        grading = self.request.context.grading
        exam = self.db.query(models.Exam).get(self.request.POST['new_exam'])
        if grading.lecture_id == exam.lecture_id:
            if not exam in grading.exams:
                grading.exams.append(exam)
                self.db.commit()
        return HTTPFound(location=self.request.route_url('grading_edit', grading_id=grading.id))

@view_config(route_name='grading_delete_exam_association', context=GradingContext, permission='edit')
class DeleteExamAssociation:
    def __init__(self, request):
        self.request = request
        self.db = self.request.db
        self.exam_id = request.matchdict['exam_id']
    def __call__(self):
        grading = self.request.context.grading
        exam = self.db.query(models.Exam).get(self.exam_id)
        if exam in grading.exams:
            grading.exams.remove(exam)
            self.db.commit()
        return HTTPFound(location=self.request.route_url('grading_edit', grading_id=grading.id))

class EnterGradesBasic:
    def __init__(self, request):
        self.request = request
        self.db = self.request.db
        self.return_json = False

    def update_grading_formula(self):
        grading = self.request.context.grading
        formula = self.request.GET.get('formula', None)
        if formula:
            grading.formula = formula
            self.db.commit()
        else:
            formula = grading.formula
        return grading, formula

    def get_lecture_students(self, grading):
        exam_id = self.request.GET.get('students', None)
        if exam_id:
            exam = self.request.db.query(models.Exam).get(exam_id)
        else: exam=None
        student_id = self.request.GET.get('student', None)
        lecture_students = grading.lecture.lecture_students_for_tutorials([]) \
            .options(sqlalchemy.orm.joinedload(models.LectureStudent.student))
        if student_id:
            lecture_students = lecture_students.filter(models.LectureStudent.student_id == student_id)
        if exam:
            lecture_students = lecture_students \
                .join(models.ExerciseStudent, models.LectureStudent.student_id==models.ExerciseStudent.student_id) \
                .join(models.Exercise) \
                .join(models.Exam) \
                .filter(models.Exam.id==exam_id)
        return lecture_students.all()

    def get_exam_vars(self, grading):
        exam_ids = [e.id for e in grading.exams]
        examvars = dict([['$%i' % i, e.id] for i,e in enumerate(grading.exams)])
        varsForExam = dict([[examvars[var], var] for var in examvars ])
        return exam_ids, examvars, varsForExam

    def get_current_grades(self, grading, lecture_students, exam_ids):
        grades = utils.autovivify()
        gradesQuery = grading.student_grades.filter(models.StudentGrade.student_id.in_([ls.student_id for ls in lecture_students]))
        for ls in lecture_students:
            grades[ls.student_id]['grade'] = ''
            grades[ls.student_id]['gradestr'] = ''
            grades[ls.student_id]['invalid'] = None
            grades[ls.student_id]['exams'] = dict([[i, {'points': '', 'admission': None, 'registration': None, 'medical_certificate': None}] for i in exam_ids])
            grades[ls.student_id]['calc'] = ''
        for grade in gradesQuery:
            grades[grade.student_id]['grade'] = grade
            grades[grade.student_id]['gradestr'] = grade.grade
        return grades

    def update_grades_with_post_params(self, grades, lecture_students, grading, error_msgs):
        if self.request.method == 'POST':
            for ls in lecture_students:
                param = 'grade-%u' % (ls.student_id)
                if param in self.request.POST:
                    value = self.request.POST[param]
                    if not grades[ls.student_id]['grade']:
                        studentGrade = models.StudentGrade()
                        studentGrade.student = ls.student
                        studentGrade.grading = grading
                        grades[ls.student_id]['grade'] = studentGrade
                        self.db.add(studentGrade)
                    if not value:
                        grades[ls.student_id]['grade'].grade = None
                        grades[ls.student_id]['gradestr'] = ''
                    else:
                        value = value.replace(',','.')
                        try:
                            value = float(value)
                            if not 0 <= value <= 6:
                                # Only accept valid grades from 0 to 6. (usually 5, but let's be lenient here)
                                raise ValueError
                            grades[ls.student_id]['grade'].grade = value
                            grades[ls.student_id]['gradestr'] = value
                        except:
                            error_msgs.append('Could not convert "%s" (%s)'%(value, ls.student.name))
        if self.db.new or self.db.dirty or self.db.deleted:
            self.db.commit()
        return grades, error_msgs

    def populate_with_exam_results(self, grades, lecture_students, grading):
        for exam in grading.exams:
            results = exam.getResults(students=lecture_students)
            for result in results:
                grades[result.student_id]['exams'][exam.id]['points'] = result.points
            if exam.admission is not None or exam.registration is not None or exam.medical_certificate is not None:
                student_ids = [ls.student_id for ls in lecture_students]
                admissions = exam.exam_admissions
                for a in admissions:
                    if a.student_id in student_ids:
                        if exam.admission is not None:
                            grades[a.student_id]['exams'][exam.id]['admission'] = a.admission
                        if exam.registration is not None:
                            grades[a.student_id]['exams'][exam.id]['registration'] = a.registration
                        if exam.medical_certificate is not None:
                            grades[a.student_id]['exams'][exam.id]['medical_certificate'] = a.medical_certificate
        return grades

    def apply_formula(self, grades, formula, lecture_students, grading, varsForExam, error_msgs):
        if formula:
            parser = Parser()
            try:
                parser.parseString(formula)
                for ls in lecture_students:
                    d = {}
                    for exam in grading.exams:
                        result = grades[ls.student_id]['exams'][exam.id]['points']
                        if result == '':
                            result = None
                        d[varsForExam[exam.id]] = result
                    result = parser.calculate(d)
                    grades[ls.student_id]['calc'] = result if result != None else ''
                    if 'fill' in self.request.GET:
                        grades[ls.student_id]['gradestr'] = result if result != None else ''
            except Exception as err:
                error_msgs.append(str(err))
        if 'fill' in self.request.GET:
            self.request.session.flash('Achtung, die Noten sind noch nicht abgespeichert!', queue='errors')
        return grades, error_msgs

    def __call__(self):
        exam_id = self.request.GET.get('students', None)
        grading, formula = self.update_grading_formula()
        lecture_students = self.get_lecture_students(grading)
        exam_ids, examvars, varsForExam = self.get_exam_vars(grading)
        grades = self.get_current_grades(grading, lecture_students, exam_ids)

        error_msgs = []
        grades, error_msgs = self.update_grades_with_post_params(grades, lecture_students, grading, error_msgs)
        grades = self.populate_with_exam_results(grades, lecture_students, grading)
        grades, error_msgs = self.apply_formula(grades, formula, lecture_students, grading, varsForExam, error_msgs)

        self.request.javascript.append('fancybox.umd.js')
        self.request.css.append('fancybox.css')
        self.request.javascript.append('toast.min.js')
        self.request.css.append('toast.min.css')
        #grades = {key: value for key,value in grades.items()}

        return {'grading': grading,
                'error_msg': '\n'.join(error_msgs),
                'formula': formula,
                'exam_id': exam_id,
                'tutorial_ids': '',
                'grades': grades,
                'examvars': examvars,
                'varsForExam': varsForExam,
                'lecture_students': lecture_students,
                'tooltips': grading_edit_tooltips}

@view_config(route_name='grading_enter_grades', renderer='muesli.web:templates/grading/enter_grades.pt', context=GradingContext, permission='edit')
class EnterGrades(EnterGradesBasic):
    pass

@view_config(route_name='grading_get_row', renderer='json', context=GradingContext, permission='edit')
class GetRow(EnterGradesBasic):
    def __init__(self, request):
        self.request = request
        self.db = self.request.db
    def __call__(self):
        result = super(GetRow, self).__call__()
        grades = result['grades']
        for key, value in list(result['grades'].items()):
            del value['grade']
            value['gradestr'] = str(value['gradestr'])
            for exam_id  in value['exams']:
                value['exams'][exam_id]['points'] = str(value['exams'][exam_id]['points'])
            if value['calc'] != None:
                value['calc'] = str(value['calc'])
        return {'grades': grades,
                'error_msg': result['error_msg']}

@view_config(route_name='grading_formula_histogram', context=GradingContext, permission='edit')
class FormulaHistogram(EnterGradesBasic):
    def generate_histogram(self):
        grading = self.request.context.grading
        formula = self.request.GET.get('formula', grading.formula)
        lecture_students = self.get_lecture_students(grading)
        exam_ids, examvars, varsForExam = self.get_exam_vars(grading)
        grades = self.get_current_grades(grading, lecture_students, exam_ids)

        error_msgs = []
        grades = self.populate_with_exam_results(grades, lecture_students, grading)
        grades, error_msgs = self.apply_formula(grades, formula, lecture_students, grading, varsForExam, error_msgs)
        grades_list = [float(grades[student_id]['calc']) for student_id in grades.keys() if not grades[student_id]['calc'] == '']

        if not grades_list:
            raise HTTPNotFound("Es sind existieren keine Noten für diese Benotung.")

        # count occurences of grades and save it in a list as tuple (grade, count)
        tuple_list = list(Counter(grades_list).items())

        # sort the list by grades
        tuple_list = sorted(tuple_list, key=lambda x: x[0])

        labels = [x[0] for x in tuple_list]
        values = [x[1] for x in tuple_list]

        indexes = numpy.arange(len(labels))
        width = 1

        pyplot.rcParams.update({'font.size': 20})

        fig = pyplot.figure(figsize=(12, 9))

        ax = fig.add_subplot(111)

        pyplot.sca(ax)
        pyplot.bar(indexes, values, width, edgecolor='black', color='red')
        pyplot.xticks(indexes + width - 1, labels)
        pyplot.xlabel('Note')
        pyplot.ylabel('Anzahl')

        yint = range(min(values), math.ceil(max(values)) + 1, math.ceil(max(values)/10))
        pyplot.yticks(yint)

        percentage_message = []

        grades_count = len(grades_list)
        percentage_list = []
        for x in range(1, 5):
            percentage_list.append(100*sum(grade <= x for grade in grades_list)/grades_count)

        percentage_message.append('• {:.1f}% haben die Note 1.0\n'.format(percentage_list[0]))
        percentage_message.append('• {:.1f}% haben die Note 2.0 oder besser\n'.format(percentage_list[1]))
        percentage_message.append('• {:.1f}% haben die Note 3.0 oder besser\n'.format(percentage_list[2]))
        percentage_message.append('• {:.1f}% haben die Note 4.0 oder besser\n'.format(percentage_list[3]))
        percentage_message.append('• {:.1f}% haben die Note 5.0\n'.format(100-percentage_list[3]))

        pyplot.text(-0.5, -3, "".join(percentage_message), verticalalignment='top')

        return fig

    def __call__(self):
        output = io.BytesIO()
        fig = self.generate_histogram()
        fig.savefig(output, format='png', dpi=50, bbox_inches='tight')
        pyplot.close(fig)
        response = Response()
        response.content_type = 'image/png'
        response.body = output.getvalue()
        output.close()
        return response

class ExcelView:
    def __init__(self, request):
        self.request = request
        self.w = Workbook()
    def createResponse(self):
        response = Response(content_type='application/vnd.ms-excel')
        with NamedTemporaryFile() as tmp:
            self.w.save(tmp.name)
            tmp.seek(0)
            response.body = tmp.read()
        return response

@view_config(route_name='grading_export', context=GradingContext, permission='edit')
class Export(ExcelView):
    def __call__(self):
        grading = self.request.context.grading
        lecture = grading.lecture
        w = self.w

        # sheet Pruefung
        worksheet_exams = w.active
        worksheet_exams.title = 'Pruefung'
        date_style = 'dd.mm.yyyy'
        header = ['Tabellenname', 'Veranstaltungsnummer', 'Titel', 'Semester', 'Termin', 'Datum', 'PrueferId', 'Pruefername']
        worksheet_exams.append(header)
        worksheet_exams.row_dimensions[1].font = Font(bold=True)
        data = ['Pruefungsteilnehmer',
                lecture.lsf_id,
                lecture.name,
                str(lecture.term) if lecture.term!=None else '',
                grading.hispos_type,
                None,
                grading.examiner_id,
                lecture.lecturer]
        worksheet_exams.append(data)
        date_p = re.compile(r"(\d{1,2}).(\d{1,2}).(\d{4})$")
        m = date_p.match(grading.hispos_date or '')
        if m:
            date = datetime.datetime(year=int(m.group(3)), month=int(m.group(2)), day=int(m.group(1)))
            date_cell = worksheet_exams.cell(row=2, column=6)
            date_cell.value = date
            date_cell.number_format = date_style
        # set column width
        for column_cells in worksheet_exams.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            worksheet_exams.column_dimensions[column_cells[0].column_letter].width = max_length*1.2

        # sheet Pruefungsteilnehmer
        worksheet_grades = self.w.create_sheet('Pruefungsteilnehmer')
        header = ['mtknr', 'name', 'stg', 'stg_txt', 'accnr', 'pnr', 'pnote', 'pstatus', 'ppunkte', 'pbonus']
        worksheet_grades.append(header)
        worksheet_grades.row_dimensions[1].font = Font(bold=True)
        grades = grading.student_grades.options(sqlalchemy.orm.joinedload(models.StudentGrade.student)).all()
        for i, grade in enumerate(grades, 1):
            if grade.grade is not None:
                g = float(grade.grade*100)
            else:
                g = ''
            data = [grade.student.matrikel,
                    grade.student.last_name, '',
                    ', '.join([s.name for s in grade.student.subjects]), '', '', g, '']
            for j, d in enumerate(data, 1):
                worksheet_grades.cell(row=1+i, column=j, value=d)
        # set column width
        for column_cells in worksheet_grades.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            worksheet_grades.column_dimensions[column_cells[0].column_letter].width = max_length*1.2
        return self.createResponse()
